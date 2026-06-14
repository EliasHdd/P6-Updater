
from __future__ import annotations

import argparse
import math
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.worksheet import Worksheet

MASTER_SHEET = "TASK"
USERDATA_SHEET = "USERDATA"

# Internal P6 export column names in row 1
MASTER_ID_INTERNAL = "task_code"
MASTER_NAME_INTERNAL = "task_name"
MASTER_ACT_START_INTERNAL = "act_start_date"
MASTER_ACT_FINISH_INTERNAL = "act_end_date"
MASTER_COMPARE_PERCENT_INTERNAL = "calc_phys_complete_pct"
MASTER_WRITE_PERCENT_INTERNAL = "user_field_212"

# Display names in row 2 for reference only
MASTER_ID_DISPLAY = "Activity ID"

LOG_HEADERS = [
    "Activity ID",
    "Activity Description",
    "Source",
    "Master Row",
    "Old %",
    "New %",
    "% Changed",
    "Actual Start",
    "Actual Finish",
    "Conflict With",
    "Result",
    "Note",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
ERR_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
GOOD_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
NOT_APPLIED_FILL = PatternFill(fill_type="solid", fgColor="FCE8C3")  # orange-ish for not-applied rows

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", NS_MAIN)


@dataclass
class UpdateRow:
    activity_id: str
    new_pct: float
    source: str
    source_sheet: str
    source_row: int
    red_font: bool = True


def clean_id(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("%", "").replace(",", ".")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_percent(value: Any) -> Optional[float]:
    num = parse_number(value)
    if num is None:
        return None
    if 0 <= num <= 1:
        num = num * 100
    return round(num, 10)


def pct_to_display(value: Optional[float]) -> Optional[int | float]:
    if value is None:
        return None
    return int(value) if float(value).is_integer() else float(value)


def is_red_font(cell) -> bool:
    color = getattr(getattr(cell, "font", None), "color", None)
    if color is None:
        return False
    if color.type == "rgb":
        rgb = (color.rgb or "").upper()
        return rgb.endswith("FF0000") or rgb in {"FFFF0000", "00FF0000"}
    return False


def parse_dt(text: str) -> datetime:
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%y %H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"Cannot parse datetime: {text}")


def find_header_row(ws: Worksheet, required_headers: List[str], max_scan_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    required = [h.casefold() for h in required_headers]
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        mapping: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str):
                mapping[value.strip().casefold()] = col
        if all(h in mapping for h in required):
            return row, {h: mapping[h.casefold()] for h in required_headers}
    raise ValueError(f"Could not find header row with columns: {required_headers} in sheet '{ws.title}'.")


def best_source_sheet(wb: openpyxl.Workbook) -> Worksheet:
    for ws in wb.worksheets:
        try:
            find_header_row(ws, ["Activity ID", "This Week's % Complete"])
            return ws
        except ValueError:
            continue
    raise ValueError("No usable subcontractor sheet found.")


def find_master_internal_headers(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    targets = [
        MASTER_ID_INTERNAL,
        MASTER_NAME_INTERNAL,
        MASTER_ACT_START_INTERNAL,
        MASTER_ACT_FINISH_INTERNAL,
        MASTER_COMPARE_PERCENT_INTERNAL,
        MASTER_WRITE_PERCENT_INTERNAL,
    ]
    for row in range(1, min(ws.max_row, 5) + 1):
        values = {
            str(ws.cell(row, c).value).strip().casefold(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row, c).value is not None
        }
        if all(t.casefold() in values for t in targets):
            return row, {t: values[t.casefold()] for t in targets}
    raise ValueError(f"Could not find expected P6 headers in sheet '{ws.title}'.")


def load_updates(source_path: Path, source_name: str) -> List[UpdateRow]:
    wb = openpyxl.load_workbook(source_path, data_only=False, keep_vba=True)
    ws = best_source_sheet(wb)
    header_row, cols = find_header_row(ws, ["Activity ID", "This Week's % Complete"])
    out: List[UpdateRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        activity_id = clean_id(ws.cell(row, cols["Activity ID"]).value)
        if not activity_id:
            continue
        pct_cell = ws.cell(row, cols["This Week's % Complete"])
        pct = normalize_percent(pct_cell.value)
        if pct is None:
            continue
        out.append(UpdateRow(
            activity_id=activity_id,
            new_pct=pct,
            source=source_name,
            source_sheet=ws.title,
            source_row=row,
            red_font=is_red_font(pct_cell),
        ))
    return out


def choose_stamp(week_finish: datetime) -> str:
    return week_finish.strftime("%Y%m%d")


def choose_output_paths(master_path: Path, week_finish: datetime) -> Tuple[Path, Path, Path]:
    stamp = choose_stamp(week_finish)
    review = master_path.with_name(f"{master_path.stem}_REVIEW_{stamp}.xlsx")
    p6_import = master_path.with_name(f"{master_path.stem}_P6_IMPORT_{stamp}.xlsx")
    log = master_path.with_name(f"{master_path.stem}_UPDATED_TASKS_LOG_{stamp}.xlsx")
    return review, p6_import, log


def ensure_sheet(wb: openpyxl.Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)
    return wb.create_sheet(title)


def autofit_columns(ws: Worksheet, max_width: int = 55) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = max(len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def _fill_for_result(result: Optional[str]) -> Optional[PatternFill]:
    if result == "NOT APPLIED":
        return NOT_APPLIED_FILL
    if result == "MASTER KEPT":
        return WARN_FILL
    if result in ("OVERRIDES",):
        return WARN_FILL
    return None


def build_log_workbook(log_rows: List[Dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Updated Tasks Log"
    ws.append(LOG_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for row in log_rows:
        ws.append([row.get(h) for h in LOG_HEADERS])
        fill = _fill_for_result(row.get("Result"))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    ws.freeze_panes = "A2"
    autofit_columns(ws)
    wb.save(output_path)


def add_log_sheet(review_wb: openpyxl.Workbook, log_rows: List[Dict[str, Any]]) -> None:
    ws = ensure_sheet(review_wb, "UPDATE_LOG")
    ws.append(LOG_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for row in log_rows:
        ws.append([row.get(h) for h in LOG_HEADERS])
        fill = _fill_for_result(row.get("Result"))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    ws.freeze_panes = "A2"
    autofit_columns(ws)


def add_conflict_sheet(review_wb: openpyxl.Workbook, conflicts: List[Dict[str, Any]]) -> None:
    ws = ensure_sheet(review_wb, "CONFLICTS")
    headers = ["Activity ID", "Activity Description", "First Source", "First %", "Winning Source", "Winning %", "Rule"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = ERR_FILL
        cell.font = Font(bold=True)
    for row in conflicts:
        ws.append([
            row.get("Activity ID"),
            row.get("Activity Description"),
            row.get("First Source"),
            row.get("First %"),
            row.get("Winning Source"),
            row.get("Winning %"),
            row.get("Rule"),
        ])
    ws.freeze_panes = "A2"
    autofit_columns(ws)


def clone_workbook(src: openpyxl.Workbook) -> openpyxl.Workbook:
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    for ws in src.worksheets:
        new_ws = new_wb.create_sheet(ws.title)
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = copy(cell._style)
                if cell.number_format:
                    new_cell.number_format = cell.number_format
                if cell.font:
                    new_cell.font = copy(cell.font)
                if cell.fill:
                    new_cell.fill = copy(cell.fill)
                if cell.border:
                    new_cell.border = copy(cell.border)
                if cell.alignment:
                    new_cell.alignment = copy(cell.alignment)
                if cell.protection:
                    new_cell.protection = copy(cell.protection)
        for col_letter, dim in ws.column_dimensions.items():
            new_wb[ws.title].column_dimensions[col_letter] = copy(dim)
        for row_idx, dim in ws.row_dimensions.items():
            new_wb[ws.title].row_dimensions[row_idx] = copy(dim)
        new_ws.freeze_panes = ws.freeze_panes
        new_ws.sheet_view.showGridLines = ws.sheet_view.showGridLines
    return new_wb


def get_style_id_for_date_column(master_path: Path, column_letter: str = "G") -> Optional[str]:
    with zipfile.ZipFile(master_path) as z:
        root = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))
        ns = {"a": NS_MAIN}
        for row_num in range(3, 200):
            cell = root.find(f".//a:sheetData/a:row[@r='{row_num}']/a:c[@r='{column_letter}{row_num}']", ns)
            if cell is not None and cell.attrib.get("s"):
                return cell.attrib["s"]
    return "1"


def patch_p6_import_copy(master_path: Path, output_path: Path, changes: Dict[int, Dict[str, Any]], cols: Dict[str, int]) -> None:
    date_style_id = get_style_id_for_date_column(master_path, "E") or "1"

    col_letters = {
        "compare_pct": get_column_letter(cols[MASTER_COMPARE_PERCENT_INTERNAL]),
        "write_pct": get_column_letter(cols[MASTER_WRITE_PERCENT_INTERNAL]),
        "act_start": get_column_letter(cols[MASTER_ACT_START_INTERNAL]),
        "act_finish": get_column_letter(cols[MASTER_ACT_FINISH_INTERNAL]),
    }

    with zipfile.ZipFile(master_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    root = ET.fromstring(data)
                    ns = {"a": NS_MAIN}
                    sheet_data = root.find("a:sheetData", ns)
                    if sheet_data is None:
                        raise ValueError("sheetData not found in TASK sheet.")
                    row_map = {int(row.attrib["r"]): row for row in sheet_data.findall("a:row", ns)}

                    for row_num, payload in changes.items():
                        row_el = row_map.get(row_num)
                        if row_el is None:
                            raise ValueError(f"Row {row_num} not found in TASK sheet XML.")

                        def get_or_create_cell(ref: str):
                            cell = row_el.find(f"a:c[@r='{ref}']", ns)
                            if cell is None:
                                cell = ET.SubElement(row_el, f"{{{NS_MAIN}}}c", {"r": ref})
                            return cell

                        # Percent as string, to match the original P6 export pattern.
                        pct_value = payload["new_pct"]
                        pct_text = str(int(pct_value) if float(pct_value).is_integer() else pct_value)
                        for pct_column in ("compare_pct", "write_pct"):
                            pct_ref = f"{col_letters[pct_column]}{row_num}"
                            pct_cell = get_or_create_cell(pct_ref)
                            pct_cell.attrib.pop("s", None)
                            pct_cell.attrib["t"] = "str"
                            v = pct_cell.find("a:v", ns)
                            if v is None:
                                v = ET.SubElement(pct_cell, f"{{{NS_MAIN}}}v")
                            v.text = pct_text

                        if payload.get("set_act_start"):
                            ref = f"{col_letters['act_start']}{row_num}"
                            c = get_or_create_cell(ref)
                            c.attrib["s"] = date_style_id
                            c.attrib["t"] = "n"
                            v = c.find("a:v", ns)
                            if v is None:
                                v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
                            v.text = str(to_excel(payload["act_start_value"]))

                        if payload.get("set_act_finish"):
                            ref = f"{col_letters['act_finish']}{row_num}"
                            c = get_or_create_cell(ref)
                            c.attrib["s"] = date_style_id
                            c.attrib["t"] = "n"
                            v = c.find("a:v", ns)
                            if v is None:
                                v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
                            v.text = str(to_excel(payload["act_finish_value"]))

                    data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                zout.writestr(item, data)


def apply_updates(master_path: Path, spie_path: Optional[Path], gcc_path: Optional[Path], week_start: datetime, week_finish: datetime,
                  review_output: Path, p6_import_output: Path, log_output: Path) -> Dict[str, Any]:
    master_wb = openpyxl.load_workbook(master_path)
    master_ws = master_wb[MASTER_SHEET]
    header_row, cols = find_master_internal_headers(master_ws)

    row_by_activity: Dict[str, int] = {}
    task_desc_by_activity: Dict[str, str] = {}
    for row in range(header_row + 2, master_ws.max_row + 1):
        activity_id = clean_id(master_ws.cell(row, cols[MASTER_ID_INTERNAL]).value)
        if activity_id:
            row_by_activity[activity_id] = row
            task_desc_by_activity[activity_id] = clean_id(master_ws.cell(row, cols[MASTER_NAME_INTERNAL]).value)

    # Use a cloned workbook for review so the original stays untouched
    review_wb = clone_workbook(master_wb)
    review_ws = review_wb[MASTER_SHEET]

    xml_changes: Dict[int, Dict[str, Any]] = {}
    log_rows: List[Dict[str, Any]] = []
    conflicts: List[Dict[str, Any]] = []
    missing_ids: List[Dict[str, Any]] = []
    non_red_changes: int = 0
    master_kept_higher: int = 0
    master_syncs: int = 0

    sources = [(p, n) for p, n in ((spie_path, "Source A"), (gcc_path, "Source B")) if p is not None]
    updates_by_activity: Dict[str, List[UpdateRow]] = {}
    for source_path, source_name in sources:
        for upd in load_updates(source_path, source_name):
            activity_id = upd.activity_id
            if activity_id not in row_by_activity:
                missing_ids.append({"id": activity_id, "source": upd.source, "pct": upd.new_pct})
                continue
            updates_by_activity.setdefault(activity_id, []).append(upd)

    for activity_id, updates in updates_by_activity.items():
        row = row_by_activity[activity_id]
        desc = task_desc_by_activity.get(activity_id, "")
        compare_pct_cell = review_ws.cell(row, cols[MASTER_COMPARE_PERCENT_INTERNAL])
        write_pct_cell = review_ws.cell(row, cols[MASTER_WRITE_PERCENT_INTERNAL])
        act_start_cell = review_ws.cell(row, cols[MASTER_ACT_START_INTERNAL])
        act_finish_cell = review_ws.cell(row, cols[MASTER_ACT_FINISH_INTERNAL])

        master_pct = normalize_percent(compare_pct_cell.value)
        if master_pct is None:
            master_pct = 0.0
        write_pct = normalize_percent(write_pct_cell.value)

        for upd in updates:
            if upd.new_pct != master_pct and not upd.red_font:
                non_red_changes += 1

        highest_source_pct = max(upd.new_pct for upd in updates)
        winning_pct = max(master_pct, highest_source_pct)
        source_values = {pct_to_display(upd.new_pct) for upd in updates}
        source_percentages_differ = len(source_values) > 1
        winner_updates = [upd for upd in updates if upd.new_pct == winning_pct]
        winning_update = winner_updates[0] if winner_updates else None

        if master_pct > highest_source_pct:
            master_kept_higher += 1
            synced_to_master = write_pct != master_pct
            if synced_to_master:
                master_syncs += 1
                compare_pct_cell.value = pct_to_display(master_pct)
                write_pct_cell.value = pct_to_display(master_pct)
                xml_changes[row] = {
                    "new_pct": master_pct,
                    "set_act_start": False,
                    "set_act_finish": False,
                    "act_start_value": None,
                    "act_finish_value": None,
                }
            conflicts.append({
                "Activity ID": activity_id,
                "Activity Description": desc,
                "First Source": "Source A/B",
                "First %": ", ".join(f"{upd.source}={pct_to_display(upd.new_pct)}" for upd in updates),
                "Winning Source": "MASTER",
                "Winning %": pct_to_display(master_pct),
                "Rule": "Master percentage higher than all contractor values; master kept",
            })
            for upd in updates:
                log_rows.append({
                    "Activity ID": activity_id,
                    "Activity Description": desc,
                    "Source": upd.source,
                    "Master Row": row,
                    "Old %": pct_to_display(master_pct),
                    "New %": pct_to_display(upd.new_pct),
                    "% Changed": "No",
                    "Actual Start": "N/A",
                    "Actual Finish": "N/A",
                    "Conflict With": "MASTER",
                    "Result": "MASTER KEPT",
                    "Note": (
                        f"Master percentage {pct_to_display(master_pct)} is higher; contractor value not imported"
                        + ("; P6 import percentage synchronized to master" if synced_to_master else "")
                    ),
                })
            continue

        if winning_pct == master_pct:
            synced_to_master = write_pct != master_pct
            if synced_to_master:
                master_syncs += 1
                compare_pct_cell.value = pct_to_display(master_pct)
                write_pct_cell.value = pct_to_display(master_pct)
                xml_changes[row] = {
                    "new_pct": master_pct,
                    "set_act_start": False,
                    "set_act_finish": False,
                    "act_start_value": None,
                    "act_finish_value": None,
                }
            if source_percentages_differ:
                conflicts.append({
                    "Activity ID": activity_id,
                    "Activity Description": desc,
                    "First Source": "Source A/B",
                    "First %": ", ".join(f"{upd.source}={pct_to_display(upd.new_pct)}" for upd in updates),
                    "Winning Source": "MASTER",
                    "Winning %": pct_to_display(master_pct),
                    "Rule": "Contractor percentages differ, but master already has the highest value",
                })
            for upd in updates:
                log_rows.append({
                    "Activity ID": activity_id,
                    "Activity Description": desc,
                    "Source": upd.source,
                    "Master Row": row,
                    "Old %": pct_to_display(master_pct),
                    "New %": pct_to_display(upd.new_pct),
                    "% Changed": "No",
                    "Actual Start": "N/A",
                    "Actual Finish": "N/A",
                    "Conflict With": "MASTER" if upd.new_pct != master_pct else "No",
                    "Result": "MASTER KEPT" if synced_to_master else "NOT APPLIED",
                    "Note": (
                        "Master already has this percentage" if upd.new_pct == master_pct else f"Lower than master; kept percentage {pct_to_display(master_pct)}"
                    ) + ("; P6 import percentage synchronized to master" if synced_to_master else ""),
                })
            continue

        if source_percentages_differ:
            conflicts.append({
                "Activity ID": activity_id,
                "Activity Description": desc,
                "First Source": "Source A/B",
                "First %": ", ".join(f"{upd.source}={pct_to_display(upd.new_pct)}" for upd in updates),
                "Winning Source": winning_update.source if winning_update else "MASTER",
                "Winning %": pct_to_display(winning_pct),
                "Rule": "Higher percentage kept",
            })

        actual_start_added = "No"
        actual_finish_added = "No"

        compare_pct_cell.value = pct_to_display(winning_pct)
        write_pct_cell.value = pct_to_display(winning_pct)

        set_act_start = False
        set_act_finish = False

        if master_pct == 0 and winning_pct > 0 and is_blank(act_start_cell.value):
            act_start_cell.value = week_start
            act_start_cell.number_format = "dd/mm/yyyy hh:mm"
            actual_start_added = "Yes"
            set_act_start = True

        if master_pct < 100 and winning_pct == 100 and is_blank(act_finish_cell.value):
            act_finish_cell.value = week_finish
            act_finish_cell.number_format = "dd/mm/yyyy hh:mm"
            actual_finish_added = "Yes"
            set_act_finish = True

        xml_changes[row] = {
            "new_pct": winning_pct,
            "set_act_start": set_act_start,
            "set_act_finish": set_act_finish,
            "act_start_value": week_start if set_act_start else None,
            "act_finish_value": week_finish if set_act_finish else None,
        }

        for upd in updates:
            is_winner = winning_update is upd
            note_parts = []
            if upd.new_pct != master_pct and not upd.red_font:
                note_parts.append("% differs from master and was considered even though cell was not marked red")
            if is_winner:
                if source_percentages_differ:
                    note_parts.append("Highest contractor percentage applied")
                result = "UPDATED"
                conflict_with = "No" if not source_percentages_differ else "Source A/B"
                actual_start = actual_start_added
                actual_finish = actual_finish_added
            elif upd.new_pct == winning_pct:
                note_parts.append(f"Matches winning percentage from {winning_update.source if winning_update else 'another source'}")
                result = "NOT APPLIED"
                conflict_with = winning_update.source if winning_update else "No"
                actual_start = "N/A"
                actual_finish = "N/A"
            else:
                note_parts.append(f"Lower than winning value; kept higher percentage {pct_to_display(winning_pct)}")
                result = "NOT APPLIED"
                conflict_with = winning_update.source if winning_update else "MASTER"
                actual_start = "N/A"
                actual_finish = "N/A"

            log_rows.append({
                "Activity ID": activity_id,
                "Activity Description": desc,
                "Source": upd.source,
                "Master Row": row,
                "Old %": pct_to_display(master_pct),
                "New %": pct_to_display(upd.new_pct),
                "% Changed": "Yes" if is_winner else "No",
                "Actual Start": actual_start,
                "Actual Finish": actual_finish,
                "Conflict With": conflict_with,
                "Result": result,
                "Note": "; ".join(note_parts) if note_parts else None,
            })

    # Save review workbook and standalone log
    add_log_sheet(review_wb, log_rows)
    add_conflict_sheet(review_wb, conflicts)
    review_wb.save(review_output)
    build_log_workbook(log_rows, log_output)

    # Build P6 import copy by patching only TASK sheet XML in the original workbook
    patch_p6_import_copy(master_path, p6_import_output, xml_changes, cols)

    applied_rows = [r for r in log_rows if r["Result"] in ("UPDATED", "OVERRIDES")]
    return {
        "review_output": str(review_output),
        "p6_import_output": str(p6_import_output),
        "log_output": str(log_output),
        "spie_updates": sum(1 for r in applied_rows if r["Source"] == "Source A"),
        "gcc_updates": sum(1 for r in applied_rows if r["Source"] == "Source B"),
        "actual_starts_added": sum(1 for r in applied_rows if r["Actual Start"] == "Yes"),
        "actual_finishes_added": sum(1 for r in applied_rows if r["Actual Finish"] == "Yes"),
        "conflicts": len(conflicts),
        "missing_ids": len(missing_ids),
        "non_red_changes": non_red_changes,
        "master_kept_higher": master_kept_higher,
        "master_syncs": master_syncs,
        "updated_tasks_logged": len(applied_rows),
        # Structured detail consumed by the web UI (server.py)
        "conflicts_detail": conflicts,
        "missing_ids_detail": missing_ids,
        "log_rows": log_rows,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Merge one or two weekly progress workbooks into a P6 export and create a P6-import copy using user_field_212 for percentages.")
    parser.add_argument("--master", required=True, help="Path to the original P6 export workbook (UP file)")
    parser.add_argument("--spie", required=False, default=None, help="Path to the Source A workbook (optional)")
    parser.add_argument("--gcc", required=False, default=None, help="Path to the Source B workbook (optional)")
    parser.add_argument("--week-start", required=True, help="Week start datetime, e.g. 16/03/2026 09:00")
    parser.add_argument("--week-finish", required=True, help="Week finish datetime, e.g. 20/03/2026 18:00")
    parser.add_argument("--review-output", help="Optional path for the review workbook")
    parser.add_argument("--p6-import-output", help="Optional path for the P6 import workbook")
    parser.add_argument("--log-output", help="Optional path for the XLSX log workbook")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    master_path = Path(args.master)
    spie_path = Path(args.spie) if args.spie else None
    gcc_path = Path(args.gcc) if args.gcc else None
    week_start = parse_dt(args.week_start)
    week_finish = parse_dt(args.week_finish)

    default_review, default_p6, default_log = choose_output_paths(master_path, week_finish)
    review_output = Path(args.review_output) if args.review_output else default_review
    p6_import_output = Path(args.p6_import_output) if args.p6_import_output else default_p6
    log_output = Path(args.log_output) if args.log_output else default_log

    summary = apply_updates(master_path, spie_path, gcc_path, week_start, week_finish, review_output, p6_import_output, log_output)

    print("Done")
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
